"""Excel eksport: qarzdorlar, ish haqi, xarajatlar, mulklar, oylik moliyaviy xulosa.
Uslub (sarlavha/qator formatlash) apps.students.export_views bilan bir xil — qayta ishlatiladi."""
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.accounts.permissions import IsFinanceOrAdmin
from apps.payments.models import Payment
from apps.students.export_views import _add_title, _response, _style_header, _style_row
from apps.teachers.models import TeacherSalaryPayment

from .models import Asset, Expense
from .services import compute_financial_summary, resolve_period


class ExportDebtsView(APIView):
    """GET /api/v1/finance/exports/debts/?month=&year="""
    permission_classes = [IsFinanceOrAdmin]

    def get(self, request):
        now = timezone.localdate()
        month = int(request.query_params.get('month', now.month))
        year  = int(request.query_params.get('year', now.year))

        debts = Payment.objects.filter(
            month=month, year=year,
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
        ).select_related('student__user', 'group').order_by('-debt_amount')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Qarzdorlar'
        ws.freeze_panes = 'A3'

        headers = ['#', "O'quvchi", 'Telefon', 'Guruh', 'Oylik to\'lov',
                   "To'langan", 'Qarz', 'Holat', 'Muddati o\'tganmi']
        _add_title(ws, f"Qarzdorlar ro'yxati — {month}/{year}", len(headers))
        _style_header(ws, headers)

        STATUS_MAP = {'paid': "To'langan", 'partial': 'Qisman', 'unpaid': "To'lanmagan"}
        for i, p in enumerate(debts, 1):
            _style_row(ws, i + 2, [
                i,
                p.student.user.full_name,
                p.student.phone,
                p.group.name if p.group else '—',
                int(p.amount),
                int(p.paid_amount),
                int(p.debt_amount),
                STATUS_MAP.get(p.status, p.status),
                'Ha' if p.is_overdue else "Yo'q",
            ], i % 2 == 0)

        col_widths = [5, 26, 16, 20, 14, 14, 14, 14, 14]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"qarzdorlar_{year}_{month:02d}.xlsx")


class ExportSalariesView(APIView):
    """GET /api/v1/finance/exports/salaries/?month=&year="""
    permission_classes = [IsFinanceOrAdmin]

    def get(self, request):
        now = timezone.localdate()
        month = int(request.query_params.get('month', now.month))
        year  = int(request.query_params.get('year', now.year))

        salaries = TeacherSalaryPayment.objects.filter(
            month=month, year=year,
        ).select_related('teacher__user', 'paid_by').order_by('teacher__user__full_name')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ish haqi'
        ws.freeze_panes = 'A3'

        headers = ['#', "O'qituvchi", 'Turi', 'Asosiy', 'Ishlagan soat', 'Bonus',
                   'Ushlab qolingan', 'Jami', 'Holat', 'Kim to\'ladi']
        _add_title(ws, f"Ish haqi to'lovlari — {month}/{year}", len(headers))
        _style_header(ws, headers)

        TYPE_MAP   = {'fixed': 'Oylik', 'hourly': 'Soatlik'}
        STATUS_MAP = {'pending': 'Kutilmoqda', 'paid': "To'langan"}
        for i, s in enumerate(salaries, 1):
            _style_row(ws, i + 2, [
                i,
                s.teacher.user.full_name,
                TYPE_MAP.get(s.teacher.salary_type, s.teacher.salary_type),
                int(s.amount),
                float(s.worked_hours) if s.worked_hours is not None else '—',
                int(s.bonus),
                int(s.deductions),
                int(s.total),
                STATUS_MAP.get(s.status, s.status),
                s.paid_by.full_name if s.paid_by else '—',
            ], i % 2 == 0)

        col_widths = [5, 26, 10, 14, 12, 12, 14, 14, 12, 18]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"ish_haqi_{year}_{month:02d}.xlsx")


class ExportExpensesView(APIView):
    """GET /api/v1/finance/exports/expenses/?month=&year= (yoki ?start=&end=)"""
    permission_classes = [IsFinanceOrAdmin]

    def get(self, request):
        try:
            start_date, end_date = resolve_period(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)

        expenses = Expense.objects.filter(
            expense_date__gte=start_date, expense_date__lte=end_date,
        ).select_related('created_by').order_by('-expense_date')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Xarajatlar'
        ws.freeze_panes = 'A3'

        headers = ['#', 'Nomi', 'Kategoriya', 'Summa', 'Sana', "Kim qo'shdi", 'Tavsif']
        _add_title(ws, f"Xarajatlar — {start_date:%d.%m.%Y} — {end_date:%d.%m.%Y}", len(headers))
        _style_header(ws, headers)

        total = 0
        for i, e in enumerate(expenses, 1):
            total += float(e.amount)
            _style_row(ws, i + 2, [
                i, e.name, e.get_category_display(), int(e.amount),
                e.expense_date.strftime('%d.%m.%Y'),
                e.created_by.full_name if e.created_by else '—',
                e.description or '—',
            ], i % 2 == 0)

        tr = expenses.count() + 3
        ws.cell(row=tr, column=1, value='JAMI').font = Font(bold=True)
        ws.cell(row=tr, column=4, value=int(total)).font = Font(bold=True, color='EF4444')

        col_widths = [5, 26, 16, 14, 14, 20, 30]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"xarajatlar_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx")


class ExportAssetsView(APIView):
    """GET /api/v1/finance/exports/assets/"""
    permission_classes = [IsFinanceOrAdmin]

    def get(self, request):
        assets = Asset.objects.select_related('created_by').order_by('name')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Mulklar'
        ws.freeze_panes = 'A3'

        headers = ['#', 'Nomi', 'Kategoriya', 'Miqdor', 'Narxi (dona)',
                   'Umumiy qiymat', 'Holati', 'Sotib olingan sana']
        _add_title(ws, f"Markaz mulklari — {timezone.now():%d.%m.%Y}", len(headers))
        _style_header(ws, headers)

        total_value = 0
        for i, a in enumerate(assets, 1):
            total_value += float(a.total_value)
            _style_row(ws, i + 2, [
                i, a.name, a.category or '—', a.quantity, int(a.purchase_price),
                int(a.total_value), a.get_condition_display(),
                a.purchase_date.strftime('%d.%m.%Y') if a.purchase_date else '—',
            ], i % 2 == 0)

        tr = assets.count() + 3
        ws.cell(row=tr, column=1, value='JAMI').font = Font(bold=True)
        ws.cell(row=tr, column=6, value=int(total_value)).font = Font(bold=True, color='059669')

        col_widths = [5, 26, 16, 10, 14, 16, 16, 16]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"mulklar_{timezone.now():%Y%m%d}.xlsx")


class ExportMonthlySummaryView(APIView):
    """GET /api/v1/finance/exports/summary/?month=&year= (yoki ?start=&end=)"""
    permission_classes = [IsFinanceOrAdmin]

    def get(self, request):
        try:
            start_date, end_date = resolve_period(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)

        data = compute_financial_summary(start_date, end_date)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Moliyaviy xulosa'

        headers = ["Ko'rsatkich", 'Summa']
        _add_title(ws, f"Moliyaviy xulosa — {start_date:%d.%m.%Y} — {end_date:%d.%m.%Y}", len(headers))
        _style_header(ws, headers)

        rows = [
            ('Umumiy daromad', int(data['total_income'])),
            ("O'qituvchilar ish haqi", int(data['total_salaries'])),
            ('Boshqa xarajatlar', int(data['total_other_expenses'])),
            ('Jami xarajat', int(data['total_expenses'])),
            ('Sof natija', int(data['net_result'])),
        ]
        for i, (label, val) in enumerate(rows, 1):
            _style_row(ws, i + 2, [label, val], i % 2 == 0)
        ws.cell(row=len(rows) + 2, column=2).font = Font(
            bold=True, color='059669' if data['net_result'] >= 0 else 'EF4444'
        )

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18

        return _response(wb, f"moliyaviy_xulosa_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx")
