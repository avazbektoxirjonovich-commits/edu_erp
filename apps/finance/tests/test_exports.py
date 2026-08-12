from io import BytesIO

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.finance.models import Asset, Expense
from apps.groups.models import Group
from apps.payments.models import Payment
from apps.students.models import Student
from apps.teachers.models import Teacher, TeacherSalaryPayment

XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


@pytest.fixture
def finance_user(db):
    return User.objects.create_user(
        phone='+998900000090', password='pass1234',
        full_name='Finance User', role=User.Role.FINANCE,
    )


@pytest.fixture
def teacher_role_user(db):
    return User.objects.create_user(
        phone='+998900000091', password='pass1234',
        full_name='Teacher User', role=User.Role.TEACHER,
    )


def auth_client(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def read_workbook(resp):
    return load_workbook(BytesIO(resp.content))


@pytest.mark.django_db
class TestExcelExports:

    def test_export_debts(self, finance_user):
        student_user = User.objects.create_user(
            phone='+998900000092', password='pass1234', full_name='Debtor', role=User.Role.STUDENT,
        )
        group = Group.objects.create(name='Export Group', start_date='2026-01-01',
                                     start_time='09:00', end_time='10:00', monthly_fee=500000)
        student = Student.objects.create(user=student_user, phone=student_user.phone, group=group)
        Payment.objects.create(student=student, group=group, month=3, year=2026, amount=500000)

        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/debts/?month=3&year=2026')
        assert resp.status_code == 200
        assert resp['Content-Type'] == XLSX_CONTENT_TYPE
        wb = read_workbook(resp)
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 'Debtor'

    def test_export_salaries(self, finance_user):
        teacher_user = User.objects.create_user(
            phone='+998900000093', password='pass1234', full_name='Salaried Teacher', role=User.Role.TEACHER,
        )
        teacher = Teacher.objects.create(user=teacher_user, phone=teacher_user.phone, salary=1000000)
        TeacherSalaryPayment.objects.create(teacher=teacher, month=3, year=2026, amount=1000000, bonus=50000)

        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/salaries/?month=3&year=2026')
        assert resp.status_code == 200
        wb = read_workbook(resp)
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 'Salaried Teacher'
        assert ws.cell(row=3, column=8).value == 1050000  # total

    def test_export_expenses(self, finance_user):
        Expense.objects.create(name='Ijara', category='rent', amount=3000000, expense_date='2026-03-05')
        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/expenses/?month=3&year=2026')
        assert resp.status_code == 200
        wb = read_workbook(resp)
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 'Ijara'

    def test_export_assets(self, finance_user):
        Asset.objects.create(name='Kompyuter', quantity=10, purchase_price=4000000)
        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/assets/')
        assert resp.status_code == 200
        wb = read_workbook(resp)
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 'Kompyuter'
        assert ws.cell(row=3, column=6).value == 40000000

    def test_export_monthly_summary(self, finance_user):
        Expense.objects.create(name='Rent', category='rent', amount=200000, expense_date='2026-03-05')
        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/summary/?month=3&year=2026')
        assert resp.status_code == 200
        wb = read_workbook(resp)
        ws = wb.active
        assert ws.cell(row=5, column=1).value == 'Boshqa xarajatlar'
        assert ws.cell(row=5, column=2).value == 200000

    def test_export_custom_date_range(self, finance_user):
        Expense.objects.create(name='Range', category='other', amount=100000, expense_date='2026-05-15')
        client = auth_client(finance_user)
        resp = client.get('/api/v1/finance/exports/expenses/?start=2026-05-01&end=2026-05-31')
        assert resp.status_code == 200

    def test_teacher_cannot_export(self, teacher_role_user):
        client = auth_client(teacher_role_user)
        for path in ['debts', 'salaries', 'expenses', 'assets', 'summary']:
            resp = client.get(f'/api/v1/finance/exports/{path}/')
            assert resp.status_code == 403, path


@pytest.mark.django_db
class TestStudentPaymentsExportOpenToFinance:

    def test_finance_can_export_payments(self, finance_user):
        client = auth_client(finance_user)
        resp = client.get('/api/v1/payments/export/?month=3&year=2026')
        assert resp.status_code == 200
        assert resp['Content-Type'] == XLSX_CONTENT_TYPE
