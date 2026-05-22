"""Excel eksport: o'quvchilar, to'lovlar, davomat"""
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.accounts.permissions import IsStaffLevel


HEADER_FILL   = PatternFill('solid', fgColor='FF6B35')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT    = Font(bold=True, size=13)
BORDER_SIDE   = Side(style='thin', color='E2E8F0')
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE, bottom=BORDER_SIDE)
ALT_FILL      = PatternFill('solid', fgColor='FFF5F2')


def _style_header(ws, headers, start_row=2):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = CELL_BORDER


def _style_row(ws, row_num, values, is_alt=False):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border    = CELL_BORDER
        if is_alt:
            cell.fill = ALT_FILL


def _add_title(ws, title, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28


def _response(wb, filename):
    res = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    res['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(res)
    return res


class ExportStudentsView(APIView):
    """GET /api/v1/students/export/"""
    permission_classes = [IsStaffLevel]

    def get(self, request):
        from .models import Student
        from django.db.models import Sum, DecimalField
        from django.db.models.functions import Coalesce

        students = (
            Student.objects
            .select_related('user', 'group')
            .annotate(
                total_debt=Coalesce(
                    Sum('payments__debt_amount',
                        output_field=DecimalField(max_digits=12, decimal_places=0)),
                    0,
                    output_field=DecimalField(max_digits=12, decimal_places=0)
                )
            )
            .order_by('user__full_name')
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"
        ws.freeze_panes = 'A3'

        headers = ['#', 'Ism Familiya', 'Telefon', 'Guruh', 'Holat',
                   "Qo'shilgan sana", 'Ota-ona', 'Ota-ona tel.', "Qarz (so'm)"]
        _add_title(ws, f"O'quvchilar ro'yxati — {timezone.now():%d.%m.%Y}", len(headers))
        _style_header(ws, headers)

        STATUS_MAP = {'active': 'Faol', 'inactive': 'Nofaol', 'frozen': 'Toxtatilgan'}
        for i, s in enumerate(students, 1):
            _style_row(ws, i + 2, [
                i,
                s.user.full_name,
                s.phone or s.user.phone or '—',
                s.group.name if s.group else '—',
                STATUS_MAP.get(s.status, s.status),
                s.joined_date.strftime('%d.%m.%Y') if s.joined_date else '—',
                s.parent_name or '—',
                s.parent_phone or '—',
                int(s.total_debt),
            ], i % 2 == 0)

        col_widths = [5, 28, 16, 20, 12, 14, 22, 16, 14]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"o_quvchilar_{timezone.now():%Y%m%d}.xlsx")


class ExportPaymentsView(APIView):
    """GET /api/v1/payments/export/?month=5&year=2025"""
    permission_classes = [IsStaffLevel]

    def get(self, request):
        from apps.payments.models import Payment
        now   = timezone.now()
        month = int(request.query_params.get('month', now.month))
        year  = int(request.query_params.get('year',  now.year))

        pays = (
            Payment.objects
            .filter(month=month, year=year)
            .select_related('student__user', 'group', 'received_by')
            .order_by('student__user__full_name')
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "To'lovlar"
        ws.freeze_panes = 'A3'

        headers = ['#', "O'quvchi", 'Guruh', 'To\'langan', 'Qarz', 'Chegirma',
                   'Holat', "To'lov sanasi", 'Qabul qildi', 'Izoh']
        _add_title(ws, f"To'lovlar — {month}/{year}", len(headers))
        _style_header(ws, headers)

        STATUS_MAP = {'paid': "To'langan", 'partial': 'Qisman', 'unpaid': "To'lanmagan"}
        total_paid = 0
        total_debt = 0
        for i, p in enumerate(pays, 1):
            total_paid += float(p.paid_amount or 0)
            total_debt += float(p.debt_amount or 0)
            _style_row(ws, i + 2, [
                i,
                p.student.user.full_name,
                p.group.name if p.group else '—',
                int(p.paid_amount or 0),
                int(p.debt_amount or 0),
                int(p.discount or 0),
                STATUS_MAP.get(p.status, p.status),
                p.payment_date.strftime('%d.%m.%Y') if p.payment_date else '—',
                p.received_by.full_name if p.received_by else '—',
                p.note or '—',
            ], i % 2 == 0)

        # Totals row
        tr = pays.count() + 3
        ws.cell(row=tr, column=1, value='JAMI').font = Font(bold=True)
        ws.cell(row=tr, column=4, value=int(total_paid)).font = Font(bold=True, color='10B981')
        ws.cell(row=tr, column=5, value=int(total_debt)).font = Font(bold=True, color='EF4444')

        col_widths = [5, 28, 20, 14, 14, 12, 14, 14, 18, 20]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"to_lovlar_{year}_{month:02d}.xlsx")


class ExportAttendanceView(APIView):
    """GET /api/v1/attendance/export/?group=<id>&month=5&year=2025"""
    permission_classes = [IsStaffLevel]

    def get(self, request):
        from apps.attendance.models import Attendance
        from apps.groups.models import Group
        now   = timezone.now()
        month = int(request.query_params.get('month', now.month))
        year  = int(request.query_params.get('year',  now.year))
        group_id = request.query_params.get('group')

        atts = (
            Attendance.objects
            .filter(date__month=month, date__year=year)
            .select_related('student__user', 'group')
            .order_by('date', 'student__user__full_name')
        )
        if group_id:
            atts = atts.filter(group_id=group_id)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Davomat'
        ws.freeze_panes = 'A3'

        headers = ['#', "O'quvchi", 'Guruh', 'Sana', 'Holat', 'Izoh']
        _add_title(ws, f'Davomat — {month}/{year}', len(headers))
        _style_header(ws, headers)

        STATUS_MAP = {'present': 'Keldi', 'absent': 'Kelmadi',
                      'late': 'Kech keldi', 'excused': 'Sababli'}
        STATUS_COLORS = {'present': '10B981', 'absent': 'EF4444',
                         'late': 'F59E0B', 'excused': '6366F1'}

        for i, a in enumerate(atts, 1):
            _style_row(ws, i + 2, [
                i,
                a.student.user.full_name,
                a.group.name if a.group else '—',
                a.date.strftime('%d.%m.%Y'),
                STATUS_MAP.get(a.status, a.status),
                a.note or '—',
            ], i % 2 == 0)
            status_cell = ws.cell(row=i + 2, column=5)
            color = STATUS_COLORS.get(a.status)
            if color:
                status_cell.font = Font(bold=True, color=color)

        col_widths = [5, 28, 20, 12, 14, 20]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        return _response(wb, f"davomat_{year}_{month:02d}.xlsx")